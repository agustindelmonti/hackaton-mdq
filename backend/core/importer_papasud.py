"""
importer_papasud.py · Importador tolerante de la planilla real de Papasud
(`Planilla_de_movimientos_2026.xls`, 12 solapas — PLAN_TRACKS_PAPASUD.md,
Track B §2).

La planilla real está sucia: nombres de columna con variaciones, celdas
vacías, fechas en formatos distintos, filas de encabezado repetidas. Este
importador NO asume que el archivo viene limpio — matchea columnas por
alias normalizado (sin acentos, sin mayúsculas, sin espacios) y devuelve,
además de las filas estructuradas, la lista de lo que no pudo interpretar
para que un humano lo revise. Nunca inventa un valor: lo que falta queda en
None con confianza 'dudosa', igual que una nota de voz.

No escribe nada solo: expone `importar(...)` que devuelve el resultado
estructurado; quien llama decide si lo persiste (regla de arquitectura #4:
nada persiste sin confirmación humana).
"""
from __future__ import annotations

import datetime
import re
import unicodedata

import openpyxl


def _norm(s: str) -> str:
    """minúsculas, sin acentos, sin espacios/puntuación — para matchear
    encabezados de columna sin importar cómo los tipeó cada persona."""
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-z0-9]", "", s.lower())
    return s


# Cada solapa declara sus columnas canónicas con los alias reales que trae la
# planilla de Papasud (ver PLAN_TRACKS_PAPASUD.md §2 · Track B).
ALIAS = {
    "remito": ["remito", "nremito", "numremito", "nrremito"],
    "fecha": ["fecha"],
    "transporte": ["transporte", "transportista"],
    "variedad": ["variedad"],
    "lote": ["lote"],
    "kg": ["kgs", "kg", "kilos"],
    "bolsas": ["bolsas", "bolsones"],
    "observaciones": ["observaciones", "observacionesdtv", "obs"],
    "valor_flete": ["valorflete", "valorfletedtv", "flete"],
    "destino": ["destino", "observacionesdestino"],
    "origen": ["origen"],
    "cliente": ["cliente"],
    "categoria": ["categoria"],
    "calibre": ["calibre"],
    "kg_prom": ["kgprom", "promedio", "prom"],
    "comisionista": ["comisionista"],
    "numero_dtvs": ["numerodtvs", "dtv", "dtvs", "nrodtv"],
    "color_bolsa": ["colorbolsa"],
    "color_hilo": ["colorhilo"],
}

# tipo_movimiento -> (nombre de campo requeridos mínimos, alias de columnas
# que esa solapa suele traer). El detector de solapa es por PALABRAS CLAVE en
# el nombre de la hoja, tolerante a variaciones ("Env a Frío", "Envio Frio",
# "ENV. A FRIO" deben matchear igual).
SOLAPAS = {
    "ingreso_tolva": {
        "match": ["ingresotolva", "tolvasantana", "ingresotolvasantana"],
        "campos": ["remito", "fecha", "transporte", "variedad", "lote", "kg",
                   "bolsas", "observaciones", "valor_flete"],
    },
    "campo_a_frio": {
        "match": ["decampoafrio", "campoafrio"],
        "campos": ["remito", "fecha", "variedad", "lote", "kg", "transporte",
                   "destino", "bolsas", "observaciones", "cliente"],
    },
    "envio_frio": {
        "match": ["envafrio", "enviofrio", "envfrio"],
        "campos": ["remito", "fecha", "variedad", "lote", "categoria", "calibre",
                   "bolsas", "kg", "transporte", "destino", "kg_prom",
                   "observaciones", "cliente"],
    },
    "retiro_frio": {
        "match": ["retfrio", "retirofrio"],
        "campos": ["fecha", "remito", "variedad", "lote", "bolsas", "transporte",
                   "origen", "destino", "kg", "kg_prom"],
    },
    "ingreso_trevelin": {
        "match": ["ingresotrevelin", "trevelin"],
        "campos": ["remito", "fecha", "transporte", "variedad", "lote", "kg",
                   "bolsas", "kg_prom", "color_bolsa", "color_hilo", "categoria"],
    },
    "entrega_cliente": {
        "match": ["entregasaclientes", "entregacliente", "entregasclientes"],
        "campos": ["remito", "fecha", "variedad", "lote", "categoria", "calibre",
                   "bolsas", "transporte", "comisionista", "destino", "kg_prom",
                   "observaciones", "numero_dtvs"],
    },
}


def _detectar_tipo(nombre_solapa: str) -> str | None:
    n = _norm(nombre_solapa)
    for tipo, spec in SOLAPAS.items():
        if any(m in n for m in spec["match"]):
            return tipo
    return None


def _mapear_encabezados(fila_encabezado: list) -> dict[str, int]:
    """Índice de columna por campo canónico, según los alias definidos."""
    normalizados = [_norm(c) for c in fila_encabezado]
    mapa: dict[str, int] = {}
    for campo, alias in ALIAS.items():
        for i, h in enumerate(normalizados):
            if h in alias:
                mapa[campo] = i
                break
    return mapa


def _parsear_fecha(valor) -> str | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime.datetime):
        return valor.date().isoformat()
    if isinstance(valor, datetime.date):
        return valor.isoformat()
    s = str(valor).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None  # fecha ilegible: se guarda como faltante, no se inventa


def _parsear_numero(valor) -> float | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _fila_vacia(fila: list) -> bool:
    return all(c is None or str(c).strip() == "" for c in fila)


def _parsear_hoja(ws, tipo: str) -> dict:
    filas_crudas = list(ws.iter_rows(values_only=True))
    filas_out = []
    errores = []

    encabezado_idx = None
    mapa = {}
    for i, fila in enumerate(filas_crudas):
        if _fila_vacia(fila):
            continue
        candidato = _mapear_encabezados(list(fila))
        # tolerante a filas de encabezado repetidas: cualquier fila que
        # matchee 'remito' y 'fecha' se toma como encabezado nuevo.
        if "remito" in candidato and "fecha" in candidato:
            encabezado_idx = i
            mapa = candidato
            continue
        if encabezado_idx is None:
            continue  # todavía no encontramos el encabezado

        registro = {"_fila_planilla": i + 1}
        campos_esperados = SOLAPAS[tipo]["campos"]
        for campo in campos_esperados:
            idx = mapa.get(campo)
            valor = fila[idx] if idx is not None and idx < len(fila) else None
            if campo == "fecha":
                registro[campo] = _parsear_fecha(valor)
            elif campo in ("kg", "bolsas", "valor_flete", "kg_prom"):
                registro[campo] = _parsear_numero(valor)
            else:
                registro[campo] = str(valor).strip() if valor not in (None, "") else None

        faltan = [c for c in ("lote", "variedad", "kg") if c in campos_esperados and not registro.get(c)]
        registro["confianza"] = "dudosa" if faltan else "confirmada"
        if faltan:
            errores.append({"fila": i + 1, "motivo": f"faltan campos: {', '.join(faltan)}"})
        filas_out.append(registro)

    return {"tipo": tipo, "filas": filas_out, "errores": errores}


def importar(ruta_archivo: str) -> dict:
    """Lee el .xlsx/.xls y devuelve un dict por solapa reconocida, más las
    solapas que no se pudieron clasificar (para que un humano las revise)."""
    wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
    resultado: dict[str, dict] = {}
    no_reconocidas: list[str] = []

    for nombre in wb.sheetnames:
        tipo = _detectar_tipo(nombre)
        if tipo is None:
            no_reconocidas.append(nombre)
            continue
        resultado[tipo] = _parsear_hoja(wb[nombre], tipo)
        resultado[tipo]["solapa_origen"] = nombre

    total_filas = sum(len(r["filas"]) for r in resultado.values())
    total_errores = sum(len(r["errores"]) for r in resultado.values())
    return {
        "solapas": resultado,
        "solapas_no_reconocidas": no_reconocidas,
        "resumen": {
            "solapas_importadas": len(resultado),
            "filas_totales": total_filas,
            "filas_con_dudas": total_errores,
        },
    }
